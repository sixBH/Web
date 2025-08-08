import time
import requests
import os
from openpyxl import Workbook
 
headers = {
    'User-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
    'referer': 'https://weibo.com/1537790411/PDQcI4AX6',
    'cookie': input('请输入你的cookies：')
}
 
 
def fetch_weibo_comments(id, uid, max_id=None):
    params = {
        'id': id,
        'is_show_bulletin': '2',
        'uid': uid,
        'fetch_level': '0',
        'locale': 'zh-CN',
    }
    if max_id:
        params['max_id'] = max_id
    response = requests.get('https://weibo.com/ajax/statuses/buildComments', params=params, headers=headers)
    data = response.json()
    return data
 
 
def write_to_excel(data, ws):
    titles = ["用户", "时间", "ID", "评论", "地区", "点赞数"]
    for col_num, title in enumerate(titles, start=1):
        ws.cell(row=1, column=col_num, value=title)
 
    for index, comment in enumerate(data, start=2):
        ws[f'A{index}'] = comment['user']['screen_name']
        ws[f'B{index}'] = comment['created_at']
        ws[f'C{index}'] = comment['id']
        ws[f'D{index}'] = comment['text_raw']
        ws[f'E{index}'] = comment['source']
        ws[f'F{index}'] = comment['like_counts']
 
 
def main():
    id = input('请输入主页id：')
    uid = input('请输入主页uid：')
    max_id_list = []
    comment_data = []
 
    for i in range(51):  # 最多爬15页
        if i == 0:
            data = fetch_weibo_comments(id, uid)
        else:
            if not max_id_list or max_id_list[-1] in ('0', '', None):
                print("没有更多评论，提前结束。")
                break
            data = fetch_weibo_comments(id, uid, max_id_list[-1])
 
        new_max_id = str(data.get('max_id', '0'))
        if new_max_id in max_id_list:
            print("max_id未更新，说明到达评论底部，提前结束。")
            break
 
        max_id_list.append(new_max_id)
        comment_data.extend(data.get('data', []))
        print(f'成功自动爬取第{i + 1}页评论')
        time.sleep(5)
 
    # 写入 Excel
    wb = Workbook()
    ws = wb.active
    write_to_excel(comment_data, ws)
    desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop', 'comment_list.xlsx')
    wb.save(desktop_path)
    print(f"评论数据已保存至 {desktop_path}")
 
 
if __name__ == '__main__':
    main()